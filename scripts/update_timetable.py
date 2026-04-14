"""타임테이블 배분안 수정 스크립트

변경 사항:
1. 블로꾸자파리 → 금요일 첫순서
2. 뽈레뽈레 → 토요일 첫순서 (이름변경)
3. 호와호 → 토요일 지누콘다 다음
4. 송인상 → 일요일 남수 다음
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

XLSX_PATH = 'docs/2026캠프 운영/2026강정피스앤뮤직캠프_타임테이블_배분안.xlsx'

# --- Color definitions ---
FILL_TRANSITION = PatternFill(start_color='FFD5D8DC', end_color='FFD5D8DC', fill_type='solid')
FILL_LARGE = PatternFill(start_color='FFFADBD8', end_color='FFFADBD8', fill_type='solid')
FILL_MEDIUM = PatternFill(start_color='FFD5F5E3', end_color='FFD5F5E3', fill_type='solid')
FILL_SOLO = {
    1: PatternFill(start_color='FFE8F6F3', end_color='FFE8F6F3', fill_type='solid'),  # 금
    2: PatternFill(start_color='FFFEF9E7', end_color='FFFEF9E7', fill_type='solid'),  # 토
    3: PatternFill(start_color='FFFDEDEC', end_color='FFFDEDEC', fill_type='solid'),  # 일
}
FILL_NONE = PatternFill(fill_type=None)

LEGEND_FILLS = [
    PatternFill(start_color='FFFADBD8', end_color='FFFADBD8', fill_type='solid'),
    PatternFill(start_color='FFD5F5E3', end_color='FFD5F5E3', fill_type='solid'),
    # 솔로/듀오 fill will be set per-day
    PatternFill(start_color='FFD5D8DC', end_color='FFD5D8DC', fill_type='solid'),
]
LEGEND_TEXTS = [
    ('■', '대형밴드 5인+ (다음 팀 전환 10분)'),
    ('■', '중형밴드 3-4인 (다음 팀 전환 7분)'),
    ('■', '솔로/듀오 (다음 팀 전환 5분)'),
    ('■', '전환시간 (다음 팀 세팅 기준)'),
]


def get_fill(scale, day):
    s = str(scale)
    if '5인' in s:
        return FILL_LARGE
    elif '3-4' in s:
        return FILL_MEDIUM
    else:
        return FILL_SOLO[day]


def get_transition_info(next_scale):
    s = str(next_scale)
    if '5인' in s:
        return 10, '대형밴드'
    elif '3-4' in s:
        return 7, '중형밴드'
    else:
        return 5, '솔로/듀오'


def time_str(minutes):
    return f"{minutes // 60}:{minutes % 60:02d}"


def extract_acts(ws):
    """시트에서 공연 데이터 추출"""
    acts = []
    for row in range(5, ws.max_row + 1):
        val_a = ws.cell(row=row, column=1).value
        if val_a is not None and isinstance(val_a, (int, float)):
            acts.append({
                'name': ws.cell(row=row, column=3).value,
                'tech': ws.cell(row=row, column=4).value,
                'scale': ws.cell(row=row, column=5).value,
                'accommodation': ws.cell(row=row, column=6).value,
                'note': ws.cell(row=row, column=7).value,
            })
    return acts


def write_sheet(ws, acts, start_time_min, day, title, num_cols=7):
    """시트에 타임테이블 데이터 쓰기"""
    # 병합 셀 모두 해제 (R5 이후)
    merges_to_remove = [r for r in ws.merged_cells.ranges if r.min_row >= 5]
    for m in merges_to_remove:
        ws.unmerge_cells(str(m))

    # 기존 데이터 영역 클리어 (R5부터)
    for row in range(5, ws.max_row + 10):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = FILL_NONE
            cell.font = Font()
            cell.alignment = Alignment()

    current = start_time_min
    row = 5

    for i, act in enumerate(acts):
        start = current
        end = start + 25
        fill = get_fill(act['scale'], day)

        # 공연 행
        data = [
            (1, float(i + 1), False),
            (2, f"{time_str(start)}~{time_str(end)}", True),
            (3, act['name'], True),
            (4, act['tech'], False),
            (5, act['scale'], False),
            (6, act['accommodation'], False),
            (7, act.get('note'), False),
        ]
        for col, val, bold in data:
            cell = ws.cell(row=row, column=col)
            cell.value = val
            cell.fill = fill
            cell.font = Font(bold=bold)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        current = end
        row += 1

        # 전환 행 (마지막 팀 제외)
        if i < len(acts) - 1:
            next_act = acts[i + 1]
            trans_min, scale_label = get_transition_info(next_act['scale'])
            trans_text = f"⟶ 전환 {trans_min}분 (다음: {next_act['name']} - {scale_label} 세팅)"

            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.fill = FILL_TRANSITION
                cell.alignment = Alignment(
                    horizontal='left' if col == 3 else 'center',
                    vertical='center'
                )
            ws.cell(row=row, column=1).value = None
            ws.cell(row=row, column=2).value = f"{time_str(current)}~{time_str(current + trans_min)}"
            ws.cell(row=row, column=3).value = trans_text

            # 전환 행 병합 (C:G)
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)

            current += trans_min
            row += 1

    end_time = current

    # R2 요약 업데이트
    large_count = sum(1 for a in acts if '5인' in str(a['scale']))
    ws.cell(row=2, column=1).value = (
        f"공연시간: {time_str(start_time_min)}~{time_str(end_time)} | "
        f"{len(acts)}팀 | 공연 25분 + 전환(다음 팀 세팅: 솔로5분/중형7분/대형10분)"
    )

    # 범례
    row += 1
    ws.cell(row=row, column=1).value = '범례:'
    row += 1
    for idx, (marker, text) in enumerate(LEGEND_TEXTS):
        if idx == 2:  # 솔로/듀오 fill (per-day)
            fill = FILL_SOLO[day]
        else:
            fill = LEGEND_FILLS[idx] if idx < len(LEGEND_FILLS) else FILL_TRANSITION
        ws.cell(row=row, column=1).value = marker
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=2).value = text
        row += 1

    return end_time, large_count


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)

    # ========== 1. 각 시트에서 현재 데이터 추출 ==========
    acts_fri = extract_acts(wb['1일차(6.5금)'])
    acts_sat = extract_acts(wb['2일차(6.6토)'])
    acts_sun = extract_acts(wb['3일차(6.7일)'])

    print("=== 변경 전 ===")
    print(f"금요일: {len(acts_fri)}팀 — {[a['name'] for a in acts_fri]}")
    print(f"토요일: {len(acts_sat)}팀 — {[a['name'] for a in acts_sat]}")
    print(f"일요일: {len(acts_sun)}팀 — {[a['name'] for a in acts_sun]}")

    # ========== 2. 데이터 수정 ==========

    # --- 금요일: 호와호 제거, 블로꾸자파리 첫순서 삽입 ---
    # 호와호 데이터 백업 (토요일로 이동할 때 사용)
    howaho = None
    for a in acts_fri:
        if '호와호' in str(a['name']):
            howaho = a.copy()
            break
    # 호와호의 테크 라이더를 PDF 기반으로 업데이트
    if howaho:
        howaho['tech'] = '마이크2, 기타앰프1, 오인페55케이블, 소형테이블, 건반스탠드'
        howaho['accommodation'] = 0.0

    acts_fri = [a for a in acts_fri if '호와호' not in str(a['name'])]

    # 블로꾸자파리 생성 (PDF #17 데이터)
    blocu = {
        'name': '블로꾸자파리',
        'tech': 'X',
        'scale': '밴드(다수)',
        'accommodation': 6.0,
        'note': '여3+남3',
    }
    acts_fri.insert(0, blocu)

    # --- 토요일: 1순서 이름변경, 호와호 삽입, 송인상 제거 ---
    # 1순서 이름 변경
    acts_sat[0]['name'] = '뽈레뽈레'

    # 송인상 데이터 백업 (일요일로 이동)
    songinsan = None
    for a in acts_sat:
        if '송인상' in str(a['name']):
            songinsan = a.copy()
            break
    acts_sat = [a for a in acts_sat if '송인상' not in str(a['name'])]

    # 호와호를 지누콘다 다음에 삽입
    jinukonda_idx = None
    for i, a in enumerate(acts_sat):
        if '지누콘다' in str(a['name']):
            jinukonda_idx = i
            break
    if jinukonda_idx is not None and howaho is not None:
        acts_sat.insert(jinukonda_idx + 1, howaho)

    # --- 일요일: 송인상을 남수 다음에 삽입 ---
    namsu_idx = None
    for i, a in enumerate(acts_sun):
        if a['name'] == '남수':
            namsu_idx = i
            break
    if namsu_idx is not None and songinsan is not None:
        acts_sun.insert(namsu_idx + 1, songinsan)

    print("\n=== 변경 후 ===")
    print(f"금요일: {len(acts_fri)}팀 — {[a['name'] for a in acts_fri]}")
    print(f"토요일: {len(acts_sat)}팀 — {[a['name'] for a in acts_sat]}")
    print(f"일요일: {len(acts_sun)}팀 — {[a['name'] for a in acts_sun]}")

    # ========== 3. 시트에 쓰기 ==========
    end_fri, large_fri = write_sheet(wb['1일차(6.5금)'], acts_fri, 18 * 60, 1, '1일차(6.5금)')
    end_sat, large_sat = write_sheet(wb['2일차(6.6토)'], acts_sat, 12 * 60, 2, '2일차(6.6토)')
    end_sun, large_sun = write_sheet(wb['3일차(6.7일)'], acts_sun, 11 * 60, 3, '3일차(6.7일)')

    print(f"\n금요일: {time_str(18*60)}~{time_str(end_fri)} ({len(acts_fri)}팀, 대형{large_fri})")
    print(f"토요일: {time_str(12*60)}~{time_str(end_sat)} ({len(acts_sat)}팀, 대형{large_sat})")
    print(f"일요일: {time_str(11*60)}~{time_str(end_sun)} ({len(acts_sun)}팀, 대형{large_sun})")

    # ========== 4. 배분요약 시트 업데이트 ==========
    ws_summary = wb['배분요약']
    # R11: 금요일
    ws_summary.cell(row=11, column=2).value = f"{time_str(18*60)}~{time_str(end_fri)}"
    ws_summary.cell(row=11, column=3).value = float(len(acts_fri))
    total_fri = end_fri - 18 * 60
    ws_summary.cell(row=11, column=4).value = f"{total_fri // 60}시간{total_fri % 60}분" if total_fri % 60 else f"{total_fri // 60}시간"
    ws_summary.cell(row=11, column=6).value = float(large_fri)

    # R12: 토요일
    ws_summary.cell(row=12, column=2).value = f"{time_str(12*60)}~{time_str(end_sat)}"
    ws_summary.cell(row=12, column=3).value = float(len(acts_sat))
    total_sat = end_sat - 12 * 60
    ws_summary.cell(row=12, column=4).value = f"{total_sat // 60}시간{total_sat % 60}분" if total_sat % 60 else f"{total_sat // 60}시간"
    ws_summary.cell(row=12, column=6).value = float(large_sat)

    # R13: 일요일
    ws_summary.cell(row=13, column=2).value = f"{time_str(11*60)}~{time_str(end_sun)}"
    ws_summary.cell(row=13, column=3).value = float(len(acts_sun))
    total_sun = end_sun - 11 * 60
    ws_summary.cell(row=13, column=4).value = f"{total_sun // 60}시간{total_sun % 60}분" if total_sun % 60 else f"{total_sun // 60}시간"
    ws_summary.cell(row=13, column=6).value = float(large_sun)

    # ========== 5. 가용일매트릭스 배정일 업데이트 ==========
    ws_matrix = wb['가용일매트릭스']
    for row in range(4, ws_matrix.max_row + 1):
        name = ws_matrix.cell(row=row, column=2).value
        if name is None:
            continue
        # 호와호: 6/5→6/6
        if '호와호' in str(name):
            ws_matrix.cell(row=row, column=7).value = '6/6(토)'
        # 송인상: 6/6→6/7
        elif '송인상' in str(name):
            ws_matrix.cell(row=row, column=7).value = '6/7(일)'
        # 블로꾸자파리/뽈레뽈레: update name and assignment
        elif '블로꾸자파리' in str(name) or '뽈레뽈레' in str(name):
            ws_matrix.cell(row=row, column=2).value = '블로꾸자파리/뽈레뽈레/동백작은학교'
            ws_matrix.cell(row=row, column=7).value = '6/5(금)+6/6(토)'

    # ========== 6. 저장 ==========
    wb.save(XLSX_PATH)
    print(f"\n저장 완료: {XLSX_PATH}")


if __name__ == '__main__':
    main()
